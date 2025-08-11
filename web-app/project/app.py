from flask import Flask, render_template, request, jsonify, Response
import cv2
import os
import numpy as np
from datetime import datetime
from ultralytics import YOLO
import face_recognition as fr
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
import threading
import time
import subprocess
import platform

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['RESULTS_FOLDER'] = 'Resultados'
app.config['PERSONAL_FOLDER'] = 'Personal'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Crear directorios si no existen
for folder in [app.config['UPLOAD_FOLDER'], app.config['RESULTS_FOLDER'], app.config['PERSONAL_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

# Variables globales para el estado de la aplicación
class AppState:
    def __init__(self):
        self.cap = None
        self.detection_active = False
        self.recognition_active = False
        self.current_source = None  # 'camera', 'image', 'video'
        self.current_frame = None
        self.epp_status = {
            'casco': False,
            'gafas': False,
            'chaleco': False,
            'persona': False,
            'guantes': False,
            'safe': False
        }
        self.camera_thread = None
        self.camera_running = False

state = AppState()

# Cargar modelo YOLO
try:
    model = YOLO('best6.pt')
    print("✅ Modelo YOLO cargado correctamente")
except Exception as e:
    print(f"❌ Error cargando modelo YOLO: {e}")
    model = None

# Cargar imágenes y nombres para reconocimiento facial
def load_face_recognition_data():
    """Carga las imágenes y nombres del directorio Personal"""
    images = []
    nombres = []
    
    if not os.path.exists(app.config['PERSONAL_FOLDER']):
        print("⚠️ Directorio Personal no encontrado")
        return [], []
    
    lista_archivos = os.listdir(app.config['PERSONAL_FOLDER'])
    
    for archivo in lista_archivos:
        try:
            img_path = os.path.join(app.config['PERSONAL_FOLDER'], archivo)
            imgdb = cv2.imread(img_path)
            if imgdb is not None:
                images.append(imgdb)
                nombres.append(os.path.splitext(archivo)[0])
        except Exception as e:
            print(f"Error cargando {archivo}: {e}")
    
    print(f"✅ Cargadas {len(images)} imágenes para reconocimiento facial")
    return images, nombres

images, nombres = load_face_recognition_data()

def codificar_rostros(images):
    """Codifica los rostros de las imágenes de referencia"""
    lista_codificaciones = []
    for img in images:
        try:
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            codificacion = fr.face_encodings(img_rgb)
            if codificacion:
                lista_codificaciones.append(codificacion[0])
        except Exception as e:
            print(f"Error codificando rostro: {e}")
    return lista_codificaciones

rostros_codificados = codificar_rostros(images)

# Mapeo de clases del modelo
class_mapping = {
    'casco': 'Casco',
    'sin casco': 'Sin Casco',
    'chaleco': 'Chaleco',
    'sin chaleco': 'Sin Chaleco',
    'gafas': 'Gafas',
    'sin gafas': 'Sin Gafas',
    'guante': 'Guantes',
    'persona': 'Persona'
}

def is_class_detected(results, class_id):
    """Verifica si una clase específica fue detectada"""
    for result in results:
        if hasattr(result, 'boxes') and result.boxes:
            for box in result.boxes:
                if int(box.cls[0].item()) == class_id:
                    return True
    return False

def realizar_reconocimiento(frame):
    """Realiza reconocimiento facial en el frame"""
    if not rostros_codificados or not nombres:
        return frame
    
    try:
        frame_pequeño = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        frame_pequeño = cv2.cvtColor(frame_pequeño, cv2.COLOR_BGR2RGB)
        
        ubicaciones_rostros = fr.face_locations(frame_pequeño, model='hog')
        codificaciones_rostros = fr.face_encodings(frame_pequeño, ubicaciones_rostros, num_jitters=1)
        
        for codificacion_rostro, ubicacion_rostro in zip(codificaciones_rostros, ubicaciones_rostros):
            coincidencias = fr.compare_faces(rostros_codificados, codificacion_rostro, tolerance=0.6)
            nombre = "Desconocido"
            
            if True in coincidencias:
                indice_coincidencia = coincidencias.index(True)
                nombre = nombres[indice_coincidencia]
                registrar_horario(nombre)
            
            # Dibujar rectángulo y nombre
            top, right, bottom, left = ubicacion_rostro
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
            
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, nombre, (left + 6, bottom - 6), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
    
    except Exception as e:
        print(f"Error en reconocimiento facial: {e}")
    
    return frame

def registrar_horario(nombre):
    """Registra la asistencia en el archivo Excel"""
    archivo_excel = "Horario.xlsx"
    
    try:
        # Crear archivo si no existe
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.append(["Nombre", "Fecha", "Hora"])
            wb.save(archivo_excel)
        
        wb = load_workbook(archivo_excel)
        ws = wb.active
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        
        # Verificar si ya se registró hoy
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0] == nombre and fila[1] == fecha_actual:
                return  # Ya registrado hoy
        
        # Registrar nueva entrada
        hora_actual = datetime.now().strftime('%H:%M:%S')
        ws.append([nombre, fecha_actual, hora_actual])
        wb.save(archivo_excel)
        print(f"✅ Registrado: {nombre} - {fecha_actual} {hora_actual}")
        
    except Exception as e:
        print(f"Error registrando horario: {e}")

def procesar_frame(frame):
    """Procesa un frame con detección de EPP y reconocimiento facial"""
    if frame is None:
        return frame
    
    frame_copy = frame.copy()
    
    # Reconocimiento facial
    if state.recognition_active:
        frame_copy = realizar_reconocimiento(frame_copy)
    
    # Detección EPP
    if state.detection_active and model:
        try:
            results = model(frame_copy, conf=0.25)
            
            # Actualizar estados EPP (solo 4 elementos)
            state.epp_status['casco'] = is_class_detected(results, 0)
            state.epp_status['chaleco'] = is_class_detected(results, 1)
            state.epp_status['gafas'] = is_class_detected(results, 2)
            state.epp_status['guantes'] = is_class_detected(results, 3)
            state.epp_status['persona'] = is_class_detected(results, 4)
            
            # Calcular cumplimiento (sin incluir persona en el cálculo)
            epp_count = sum([state.epp_status['casco'], state.epp_status['chaleco'], 
                           state.epp_status['gafas'], state.epp_status['guantes']])
            state.epp_status['safe'] = epp_count == 4
            
            # Dibujar detecciones
            frame_copy = dibujar_detecciones(frame_copy, results)
            
        except Exception as e:
            print(f"Error en detección EPP: {e}")
    
    return frame_copy

def dibujar_detecciones(frame, results):
    """Dibuja las cajas de detección en el frame"""
    try:
        for result in results:
            if hasattr(result, 'boxes') and result.boxes:
                for box in result.boxes:
                    x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                    class_id = int(box.cls[0].item())
                    confidence = box.conf[0].item()
                    
                    if class_id < len(model.names):
                        class_name = model.names[class_id]
                        label = class_mapping.get(class_name, "Desconocido")
                        
                        # Color según el tipo de detección
                        if 'sin' in class_name.lower():
                            color = (0, 0, 255)  # Rojo para elementos faltantes
                        else:
                            color = (0, 255, 0)  # Verde para elementos detectados
                        
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                        cv2.putText(frame, f"{label} ({confidence:.2f})", 
                                   (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
    except Exception as e:
        print(f"Error dibujando detecciones: {e}")
    
    return frame

def camera_worker():
    """Hilo para manejar la cámara"""
    while state.camera_running and state.cap and state.cap.isOpened():
        try:
            ret, frame = state.cap.read()
            if ret:
                # Redimensionar frame para mejor performance
                frame = cv2.resize(frame, (640, 480))
                state.current_frame = procesar_frame(frame)
            else:
                break
        except Exception as e:
            print(f"Error en camera_worker: {e}")
            break
        
        time.sleep(0.033)  # ~30 FPS

def generate_frames():
    """Generador de frames para el stream de video"""
    while True:
        if state.current_frame is not None:
            try:
                # Codificar frame como JPEG
                ret, buffer = cv2.imencode('.jpg', state.current_frame, 
                                         [cv2.IMWRITE_JPEG_QUALITY, 85])
                if ret:
                    frame_bytes = buffer.tobytes()
                    yield (b'--frame\r\n'
                           b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
            except Exception as e:
                print(f"Error generando frame: {e}")
        
        time.sleep(0.033)  # ~30 FPS

# ====== RUTAS FLASK ======

@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    """Stream de video en tiempo real"""
    return Response(generate_frames(),
                   mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/upload_file', methods=['POST'])
def upload_file():
    """Subir archivo de imagen o video"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Determinar tipo de archivo
        file_type = 'image' if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')) else 'video'
        
        # Limpiar estado anterior
        state.camera_running = False
        if state.camera_thread and state.camera_thread.is_alive():
            state.camera_thread.join(timeout=1)
        
        if state.cap:
            state.cap.release()
            state.cap = None
        
        if file_type == 'image':
            # Cargar imagen
            img = cv2.imread(filepath)
            if img is not None:
                img = cv2.resize(img, (640, 480))
                state.current_frame = procesar_frame(img)
                state.current_source = 'image'
                print(f"✅ Imagen cargada: {filename}")
            else:
                return jsonify({'success': False, 'error': 'Invalid image file'})
        else:
            # Cargar video
            state.cap = cv2.VideoCapture(filepath)
            if state.cap.isOpened():
                state.current_source = 'video'
                state.camera_running = True
                state.camera_thread = threading.Thread(target=camera_worker)
                state.camera_thread.daemon = True
                state.camera_thread.start()
                print(f"✅ Video cargado: {filename}")
            else:
                return jsonify({'success': False, 'error': 'Invalid video file'})
        
        return jsonify({
            'success': True, 
            'type': file_type,
            'filename': filename,
            'url': f'/uploads/{filename}' if file_type == 'image' else None
        })
        
    except Exception as e:
        print(f"Error uploading file: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Servir archivos subidos"""
    try:
        from flask import send_from_directory
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        print(f"Error serving file: {e}")
        return "File not found", 404

@app.route('/start_camera', methods=['POST'])
def start_camera():
    """Activar cámara"""
    try:
        # Limpiar estado anterior
        if state.cap:
            state.cap.release()
            state.cap = None
        
        # Intentar abrir cámara (probar índices 0, 1, 2)
        for camera_index in [0, 1, 2]:
            state.cap = cv2.VideoCapture(camera_index)
            if state.cap.isOpened():
                # Configurar cámara
                state.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                state.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                state.cap.set(cv2.CAP_PROP_FPS, 30)
                
                state.current_source = 'camera'
                state.camera_running = True
                
                # Iniciar hilo de cámara
                if state.camera_thread and state.camera_thread.is_alive():
                    state.camera_running = False
                    state.camera_thread.join()
                
                state.camera_thread = threading.Thread(target=camera_worker)
                state.camera_thread.daemon = True
                state.camera_thread.start()
                
                print(f"✅ Cámara activada en índice {camera_index}")
                return jsonify({'success': True, 'camera_index': camera_index})
        
        return jsonify({'success': False, 'error': 'No se pudo acceder a ninguna cámara'})
        
    except Exception as e:
        print(f"Error starting camera: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/stop_camera', methods=['POST'])
def stop_camera():
    """Desactivar cámara"""
    try:
        state.camera_running = False
        
        if state.camera_thread and state.camera_thread.is_alive():
            state.camera_thread.join(timeout=2)
        
        if state.cap:
            state.cap.release()
            state.cap = None
        
        state.current_source = None
        state.current_frame = None
        
        print("✅ Cámara desactivada")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Error stopping camera: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/toggle_detection', methods=['POST'])
def toggle_detection():
    """Activar/desactivar detección EPP"""
    try:
        state.detection_active = not state.detection_active
        
        # Si se desactiva, resetear estados EPP
        if not state.detection_active:
            state.epp_status = {key: False for key in state.epp_status}
        else:
            # Si se activa y hay una imagen cargada, reprocesar
            if state.current_source == 'image' and state.current_frame is not None:
                # Recargar y procesar la imagen original
                upload_folder = app.config['UPLOAD_FOLDER']
                image_files = [f for f in os.listdir(upload_folder) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp'))]
                if image_files:
                    # Tomar la imagen más reciente
                    latest_image = max(image_files, key=lambda x: os.path.getctime(os.path.join(upload_folder, x)))
                    img_path = os.path.join(upload_folder, latest_image)
                    img = cv2.imread(img_path)
                    if img is not None:
                        img = cv2.resize(img, (640, 480))
                        state.current_frame = procesar_frame(img)
        
        print(f"✅ Detección EPP {'activada' if state.detection_active else 'desactivada'}")
        return jsonify({'success': True, 'active': state.detection_active})
        
    except Exception as e:
        print(f"Error toggling detection: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/toggle_recognition', methods=['POST'])
def toggle_recognition():
    """Activar/desactivar reconocimiento facial"""
    try:
        state.recognition_active = not state.recognition_active
        
        print(f"✅ Reconocimiento facial {'activado' if state.recognition_active else 'desactivado'}")
        return jsonify({'success': True, 'active': state.recognition_active})
        
    except Exception as e:
        print(f"Error toggling recognition: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/capture_frame', methods=['POST'])
def capture_frame():
    """Capturar frame actual"""
    try:
        if state.current_frame is None:
            return jsonify({'success': False, 'error': 'No hay frame activo para capturar'})
        
        # Generar nombre único basado en timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Captura_{timestamp}.jpg"
        filepath = os.path.join(app.config['RESULTS_FOLDER'], filename)
        
        # Guardar frame actual
        cv2.imwrite(filepath, state.current_frame)
        
        print(f"✅ Captura guardada: {filename}")
        return jsonify({'success': True, 'filename': filename})
        
    except Exception as e:
        print(f"Error capturing frame: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_detection_status')
def get_detection_status():
    """Obtener estado actual de detección EPP"""
    return jsonify(state.epp_status)

@app.route('/get_initial_state')
def get_initial_state():
    """Obtener estado inicial de la aplicación"""
    return jsonify({
        'detection_active': state.detection_active,
        'recognition_active': state.recognition_active,
        'current_source': state.current_source,
        'epp_status': state.epp_status
    })

@app.route('/reset_system', methods=['POST'])
def reset_system():
    """Resetear todo el sistema"""
    try:
        # Detener cámara
        state.camera_running = False
        if state.camera_thread and state.camera_thread.is_alive():
            state.camera_thread.join(timeout=2)
        
        if state.cap:
            state.cap.release()
            state.cap = None
        
        # Resetear estados
        state.detection_active = False
        state.recognition_active = False
        state.current_source = None
        state.current_frame = None
        state.epp_status = {key: False for key in state.epp_status}
        
        print("✅ Sistema reiniciado")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Error resetting system: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/open_excel')
def open_excel():
    """Abrir archivo Excel de horarios"""
    try:
        excel_file = "Horario.xlsx"
        
        # Crear archivo si no existe
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Nombre", "Fecha", "Hora"])
            wb.save(excel_file)
        
        # Abrir archivo según el sistema operativo
        system = platform.system()
        if system == "Windows":
            os.startfile(excel_file)
        elif system == "Darwin":  # macOS
            subprocess.call(["open", excel_file])
        else:  # Linux
            subprocess.call(["xdg-open", excel_file])
        
        print(f"✅ Archivo Excel abierto: {excel_file}")
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Error opening Excel: {e}")
        return jsonify({'success': False, 'error': str(e)})

# ====== MANEJO DE ERRORES ======

@app.errorhandler(404)
def not_found(error):
    return jsonify({'success': False, 'error': 'Endpoint no encontrado'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500

@app.errorhandler(413)
def file_too_large(error):
    return jsonify({'success': False, 'error': 'Archivo demasiado grande (máximo 50MB)'}), 413

# ====== LIMPIEZA AL CERRAR ======

import atexit

def cleanup():
    """Limpieza al cerrar la aplicación"""
    print("🧹 Limpiando recursos...")
    state.camera_running = False
    if state.cap:
        state.cap.release()
    if state.camera_thread and state.camera_thread.is_alive():
        state.camera_thread.join(timeout=2)

atexit.register(cleanup)

# ====== CONFIGURACIÓN ADICIONAL ======

# Configurar logging
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Deshabilitar logs de werkzeug para requests del video feed
logging.getLogger('werkzeug').setLevel(logging.ERROR)

if __name__ == '__main__':
    print("🚀 Iniciando Sistema de Detección EPP...")
    print("📁 Estructura de directorios creada")
    print("🔍 Modelo YOLO cargado" if model else "⚠️ Modelo YOLO no disponible")
    print(f"👥 {len(rostros_codificados)} rostros codificados para reconocimiento")
    print("🌐 Servidor iniciando en http://localhost:5000")
    
    # Iniciar servidor Flask
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=False,  # Cambiar a True para desarrollo
        threaded=True,
        use_reloader=False  # Evitar problemas con threads
    )